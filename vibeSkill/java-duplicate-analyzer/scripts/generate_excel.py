#!/usr/bin/env python3
"""
Excel 报告生成脚本

根据 jscpd 的 JSON 报告生成 Excel 表格，包含所有需要整改的代码重复项。

Usage:
    python generate_excel.py <jscpd-report.json> <output-excel.xlsx>
    
Example:
    python generate_excel.py .jscpd-report/jscpd-report.json ./duplicate-report.xlsx
"""

import sys
import json
from pathlib import Path
from datetime import datetime


def extract_filename(filepath):
    """从完整路径中提取文件名"""
    # 处理 dict 类型的 firstFile（如 jscpd 新版本）
    if isinstance(filepath, dict):
        filepath = filepath.get('name', '')
    if not filepath:
        return ""
    # 处理反斜杠和正斜杠
    filepath = str(filepath).replace("\\", "/")
    return filepath.split("/")[-1]


def extract_path(filepath):
    """从完整路径中提取相对路径（去掉盘符和前缀）"""
    # 处理 dict 类型的 firstFile（如 jscpd 新版本）
    if isinstance(filepath, dict):
        filepath = filepath.get('name', '')
    if not filepath:
        return ""
    # 处理反斜杠
    filepath = str(filepath).replace("\\", "/")
    return filepath


def get_line_info(dup, key):
    """获取行号信息，兼容新旧版本"""
    file_info = dup.get(key, {})
    if isinstance(file_info, dict):
        start = file_info.get('start', 0)
        end = file_info.get('end', 0)
    else:
        start = dup.get(f'{key}Start', 0)
        end = dup.get(f'{key}End', 0)
    return start, end


def generate_excel(json_path, output_path):
    """根据 JSON 报告生成 Excel 文件"""
    
    # 检查输入文件
    json_file = Path(json_path)
    if not json_file.exists():
        print(f"Error: JSON report not found: {json_path}")
        return False
    
    # 读取 JSON 数据
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error reading JSON file: {e}")
        return False
    
    # 尝试导入 openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # 创建工作簿
    wb = Workbook()
    
    # ===== Sheet 1: 概览统计 =====
    ws_summary = wb.active
    ws_summary.title = "概览统计"
    
    # 获取统计数据
    stats = data.get('statistics', {})
    duplicates = data.get('duplicates', [])
    
    # 计算总重复行数
    total_lines = sum(dup.get('lines', 0) for dup in duplicates)
    
    # 设置列宽
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    
    # 标题样式
    title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入标题
    ws_summary['A1'] = "Java 代码重复分析报告"
    ws_summary['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = "指标"
    ws_summary['B3'] = "数值"
    for cell in ['A3', 'B3']:
        ws_summary[cell].font = Font(name='微软雅黑', bold=True)
        ws_summary[cell].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws_summary[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入统计数据
    summary_data = [
        ("扫描时间", stats.get('detectionDate', 'N/A')),
        ("重复块数量", f"{len(duplicates)} 个"),
        ("总重复行数", f"{total_lines:,} 行"),
        ("扫描范围", "hss-domain/src 目录"),
    ]
    
    for i, (key, value) in enumerate(summary_data, start=4):
        ws_summary[f'A{i}'] = key
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = Font(name='微软雅黑')
        ws_summary[f'B{i}'].font = Font(name='微软雅黑')
    
    # ===== Sheet 2: 重复代码明细 =====
    ws_detail = wb.create_sheet("重复代码明细")
    
    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = [
        "序号",
        "第一个文件",
        "第二个文件",
        "第一个文件行号",
        "第二个文件行号",
        "重复行数",
        "严重程度",
        "重构建议"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_detail.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    ws_detail.column_dimensions['A'].width = 8
    ws_detail.column_dimensions['B'].width = 60
    ws_detail.column_dimensions['C'].width = 60
    ws_detail.column_dimensions['D'].width = 20
    ws_detail.column_dimensions['E'].width = 20
    ws_detail.column_dimensions['F'].width = 12
    ws_detail.column_dimensions['G'].width = 12
    ws_detail.column_dimensions['H'].width = 30
    
    # 按重复行数排序
    sorted_duplicates = sorted(duplicates, key=lambda x: x.get('lines', 0), reverse=True)
    
    # 写入数据
    for row, dup in enumerate(sorted_duplicates, start=2):
        first_file = dup.get('firstFile', '')
        second_file = dup.get('secondFile', '')
        lines = dup.get('lines', 0)
        
        # 解析行号 - 使用新的兼容函数
        first_start, first_end = get_line_info(dup, 'firstFile')
        second_start, second_end = get_line_info(dup, 'secondFile')
        
        # 确定严重程度
        if lines >= 100:
            severity = "严重"
        elif lines >= 50:
            severity = "较严重"
        elif lines >= 20:
            severity = "一般"
        else:
            severity = "轻微"
        
        # 重构建议
        if lines >= 100:
            refactor = "建议提取公共基类或工具类"
        elif lines >= 50:
            refactor = "建议抽取公共方法"
        else:
            refactor = "保持观察"
        
        # 写入数据
        ws_detail.cell(row=row, column=1).value = row - 1  # 序号
        ws_detail.cell(row=row, column=2).value = extract_path(first_file)
        ws_detail.cell(row=row, column=3).value = extract_path(second_file)
        ws_detail.cell(row=row, column=4).value = f"{first_start}-{first_end}"
        ws_detail.cell(row=row, column=5).value = f"{second_start}-{second_end}"
        ws_detail.cell(row=row, column=6).value = lines
        ws_detail.cell(row=row, column=7).value = severity
        ws_detail.cell(row=row, column=8).value = refactor
        
        # 设置单元格样式
        for col in range(1, 9):
            cell = ws_detail.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='微软雅黑')
            
            # 严重程度颜色
            if col == 7:
                if severity == "严重":
                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    cell.font = Font(name='微软雅黑', color='FFFFFF', bold=True)
                elif severity == "较严重":
                    cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                elif severity == "一般":
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    
    # ===== Sheet 3: Top 10 严重重复 =====
    ws_top = wb.create_sheet("Top 10 严重重复")
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = ws_top.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for col in range(1, 9):
        ws_top.column_dimensions[chr(64 + col)].width = ws_detail.column_dimensions[chr(64 + col)].width
    
    # 取 Top 10
    top_duplicates = sorted_duplicates[:10]
    
    for row, dup in enumerate(top_duplicates, start=2):
        first_file = dup.get('firstFile', '')
        second_file = dup.get('secondFile', '')
        lines = dup.get('lines', 0)
        first_start, first_end = get_line_info(dup, 'firstFile')
        second_start, second_end = get_line_info(dup, 'secondFile')
        
        severity = "严重" if lines >= 100 else "较严重" if lines >= 50 else "一般"
        refactor = "建议提取公共基类或工具类" if lines >= 100 else "建议抽取公共方法" if lines >= 50 else "保持观察"
        
        ws_top.cell(row=row, column=1).value = row - 1
        ws_top.cell(row=row, column=2).value = extract_path(first_file)
        ws_top.cell(row=row, column=3).value = extract_path(second_file)
        ws_top.cell(row=row, column=4).value = f"{first_start}-{first_end}"
        ws_top.cell(row=row, column=5).value = f"{second_start}-{second_end}"
        ws_top.cell(row=row, column=6).value = lines
        ws_top.cell(row=row, column=7).value = severity
        ws_top.cell(row=row, column=8).value = refactor
        
        for col in range(1, 9):
            cell = ws_top.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='微软雅黑')
            if col == 7:
                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                cell.font = Font(name='微软雅黑', color='FFFFFF', bold=True)
    
    # 保存文件
    try:
        wb.save(output_path)
        print(f"Excel report generated: {output_path}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_excel.py <jscpd-report.json> <output-excel.xlsx>")
        print("Example: python generate_excel.py .jscpd-report/jscpd-report.json ./duplicate-report.xlsx")
        sys.exit(1)
    
    json_path = sys.argv[1]
    output_path = sys.argv[2]
    
    success = generate_excel(json_path, output_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()