#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Java Code Scanner v2 - 核心扫描引擎
====================================
功能:
  1. 冗余重复代码扫描（基于 jscpd）
  2. 无用代码扫描（基于 javalang AST 解析器，检测未使用的 import/字段/方法/局部变量）
  3. 生成 Excel 报告 + Markdown 报告

用法:
  python main.py --project-path /path/to/java/code [--output report.md|xlsx]
  python main.py --project-path /path --min-lines 5 --min-tokens 100
  python main.py --project-path /path --skip-jscpd
  python main.py --project-path /path --skip-dead-code
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

# ── Excel 依赖（可选）──────────────────────────────────────────────
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ====================================================================
# 常量
# ====================================================================

PRIORITY_HIGH = "高"
PRIORITY_MEDIUM = "中"
PRIORITY_LOW = "低"

PRIORITY_ORDER = {PRIORITY_HIGH: 0, PRIORITY_MEDIUM: 1, PRIORITY_LOW: 2}

CST = timezone(timedelta(hours=8))

HEADER_DUPLICATE = [
    "重要程度", "项目相对路径", "文件名", "起始行", "结束行",
    "问题类型", "详细说明（重复代码位置）",
]

HEADER_DEAD_CODE = [
    "重要程度", "项目相对路径", "文件名", "起始行", "结束行",
    "问题类型", "详细说明（为什么无用）",
]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
CELL_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 高/中/低行颜色
PRIORITY_FILLS = {
    PRIORITY_HIGH: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    PRIORITY_MEDIUM: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    PRIORITY_LOW: None,
}

# 文件发现配置
JAVA_PATTERN = "**/*.java"
EXCLUDE_DIRS = {".git", "node_modules", "target", "build", "out", ".idea", ".vscode", "__pycache__",
                ".jscpd-report", "jscpd-report", "dist", "classes"}


# ====================================================================
# 工具函数
# ====================================================================

def info(msg: str) -> None:
    print(f"  [INFO] {msg}")

def ok(msg: str) -> None:
    print(f"  [OK]   {msg}")

def warn(msg: str) -> None:
    print(f"  [WARN] {msg}")

def err(msg: str) -> None:
    print(f"  [ERR]  {msg}")

def skip(msg: str) -> None:
    print(f"  [SKIP] {msg}")


def resolve_path(path_str: str) -> Path:
    """规范化并验证路径。"""
    p = Path(path_str).resolve()
    if not p.exists():
        raise FileNotFoundError(f"路径不存在: {path_str}")
    if not p.is_dir():
        raise NotADirectoryError(f"路径不是目录: {path_str}")
    return p


def rel_path(project_root: Path, abs_path: str) -> str:
    """获取文件相对于 project_root 的相对路径。"""
    try:
        return str(Path(abs_path).resolve().relative_to(project_root))
    except (ValueError, OSError):
        return Path(abs_path).name


def find_java_dirs(project_root: Path) -> List[str]:
    """
    在目录树中递归寻找所有包含 .java 文件的子目录。
    如果 project_root 本身有 .java 文件，返回 ['.']。
    """
    java_subdirs: Set[str] = set()
    root_has_java = False

    for dirpath, dirnames, filenames in os.walk(str(project_root)):
        # 跳过排除目录
        dirnames[:] = [d for d in dirnames if d not in EXCLUDE_DIRS]

        has_java = any(f.endswith(".java") for f in filenames)
        if has_java:
            if dirpath == str(project_root):
                root_has_java = True
            else:
                rel = rel_path(project_root, dirpath)
                if rel and rel != ".":
                    java_subdirs.add(rel)

    if root_has_java:
        return ["."]

    return sorted(java_subdirs)


# ====================================================================
# 环境检查
# ====================================================================

def check_environment() -> Dict[str, bool]:
    """检查所有运行依赖，缺失时只做记录。"""
    print("\n" + "=" * 60)
    info("环境依赖检查")
    print("=" * 60)

    status: Dict[str, bool] = {}

    # Python
    ver = sys.version_info
    py_ok = ver.major >= 3 and ver.minor >= 9
    ok(f"Python {ver.major}.{ver.minor}.{ver.micro}") if py_ok else err("需要 Python >= 3.9")
    status["python"] = py_ok

    # jscpd (直接使用 shutil.which)
    jscpd_bin = shutil.which("jscpd")
    if jscpd_bin:
        ok(f"jscpd ({jscpd_bin})")
        status["jscpd"] = True
    else:
        warn("jscpd 未安装 (将跳过冗余代码扫描)")
        status["jscpd"] = False

    # javalang
    try:
        import javalang  # noqa: F401
        ok("javalang 已安装")
        status["javalang"] = True
    except ImportError:
        warn("javalang 未安装 (将跳过无用代码扫描)")
        status["javalang"] = False

    # pandas + openpyxl
    if HAS_PANDAS:
        ok("pandas 已安装")
    else:
        warn("pandas 未安装 (将跳过 Excel 报告)")

    if HAS_OPENPYXL:
        ok("openpyxl 已安装")
    else:
        warn("openpyxl 未安装 (将跳过 Excel 报告)")

    print()
    return status


# ====================================================================
# 冗余重复代码扫描 (jscpd)
# ====================================================================

def _get_start_line(file_info: Dict[str, Any]) -> int:
    """
    从 jscpd report 的 file info 中提取起始行号。
    兼容 jscpd 3.x / 4.x 以及直接整数或嵌套对象格式。
    """
    val = file_info.get("start", 0)
    if isinstance(val, int) and val > 0:
        return val
    if isinstance(val, dict):
        return val.get("line", 0)
    return file_info.get("startLoc", {}).get("line", 0)


def _get_end_line(file_info: Dict[str, Any]) -> int:
    """提取结束行号，兼容多种 jscpd 版本。"""
    val = file_info.get("end", 0)
    if isinstance(val, int) and val > 0:
        return val
    if isinstance(val, dict):
        return val.get("line", 0)
    return file_info.get("endLoc", {}).get("line", 0)


def _find_jscpd_report(report_dir: str) -> Optional[str]:
    """
    在 report_dir 中查找 jscpd 生成的 JSON 报告文件。
    排除用作 jscpd 配置的 .jscpd.json 文件。
    """
    for dirpath, _, filenames in os.walk(report_dir):
        for f in sorted(filenames):
            full = os.path.join(dirpath, f)
            basename = os.path.basename(full)
            if not f.endswith(".json"):
                continue
            # 跳过我们自己写的配置文件 (jscpd 默认读 .jscpd.json 作为配置)
            if basename == ".jscpd.json":
                continue
            if "jscpd" in basename.lower():
                return full
    # fallback: 找任意 json
    for dirpath, _, filenames in os.walk(report_dir):
        for f in sorted(filenames):
            full = os.path.join(dirpath, f)
            if f.endswith(".json") and os.path.basename(full) != ".jscpd.json":
                return full
    return None


def scan_duplicate(
    scan_target: Path,
    temp_dir: str,
    project_root: Path,
    min_lines: int = 3,
    min_tokens: int = 50,
) -> Tuple[List[Dict[str, Any]], List[str], int, int]:
    """
    使用 jscpd 扫描 Java 文件的重复代码。

    Args:
        scan_target: 扫描目标目录
        temp_dir: 临时目录用于存放 jscpd 输出
        project_root: 项目根目录
        min_lines: 最小重复行数
        min_tokens: 最小 token 数

    Returns:
        (结果列表, 错误列表, 总文件数, 总重复块数)
    """
    print("\n" + "=" * 60)
    info("冗余重复代码扫描 - jscpd")
    print("=" * 60)

    results: List[Dict[str, Any]] = []
    errors: List[str] = []
    total_files = 0
    total_clones = 0

    jscpd_bin = shutil.which("jscpd")
    if not jscpd_bin:
        return results, ["jscpd 未安装"], 0, 0

    info(f"jscpd: {jscpd_bin}")
    info(f"扫描目标: {scan_target}")
    info(f"参数: min-lines={min_lines}, min-tokens={min_tokens}")

    # 使用显式配置文件传递参数，避免同名文件冲突
    # 关键：配置文件命名为 scanner-config.json（不是 .jscpd.json）
    config_path = os.path.join(temp_dir, "scanner-config.json")
    config: Dict[str, Any] = {
        "path": [str(scan_target)],
        "pattern": JAVA_PATTERN,
        "reporters": ["json"],
        "output": temp_dir,
        "threshold": 0,
        "minLines": min_lines,
        "minTokens": min_tokens,
        "silent": True,
    }

    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)

    cmd = [jscpd_bin, "--config", config_path]

    try:
        result = subprocess.run(
            cmd,
            cwd=str(project_root),
            capture_output=True,
            text=True,
            timeout=600,
        )
    except subprocess.TimeoutExpired:
        errors.append("jscpd 扫描超时（超过 600 秒）")
        return results, errors, 0, 0

    stderr_text = (result.stderr or "").strip()
    stdout_text = (result.stdout or "").strip()

    # 打印 jscpd 摘要输出
    if stdout_text:
        for line in stdout_text.split("\n")[:10]:
            if line.strip():
                print(f"       | {line.strip()}")

    # jscpd 发现重复后可能非零退出，先尝试找报告
    if result.returncode != 0:
        if "Too many duplicates" in stderr_text or "over threshold" in stderr_text:
            warn("jscpd 检测到重复超过阈值，继续解析报告")
        elif "No files found" in stderr_text or "no files" in stderr_text.lower():
            errors.append(f"jscpd 未找到匹配的 Java 文件: {stderr_text[:300]}")
            return results, errors, 0, 0
        else:
            warn(f"jscpd 退出码 {result.returncode}，尝试解析报告")

    # 查找报告文件
    report_path = _find_jscpd_report(temp_dir)

    if not report_path:
        info("jscpd 未生成报告，未发现重复代码")
        return results, errors, 0, 0

    info(f"解析报告: {os.path.basename(report_path)}")

    try:
        with open(report_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as je:
        errors.append(f"jscpd 报告 JSON 解析失败: {je}")
        return results, errors, 0, 0

    # 提取统计信息
    stats = data.get("statistics", {})
    total_stats = stats.get("total", {})
    total_files = total_stats.get("sources", 0)
    total_clones = total_stats.get("clones", 0)

    duplicates = data.get("duplicates", [])
    if not duplicates:
        info(f"未发现重复代码 (扫描 {total_files} 个文件)")
        return results, errors, total_files, total_clones

    info(f"发现 {len(duplicates)} 处重复 (扫描 {total_files} 个文件)")

    for dup in duplicates:
        try:
            # jscpd 4.x 使用 firstFile/secondFile
            first = dup.get("firstFile", {})
            second = dup.get("secondFile", {})

            f_file = first.get("name", "") or first.get("path", "")
            s_file = second.get("name", "") or second.get("path", "")

            if not f_file or not s_file:
                continue

            f_start = _get_start_line(first)
            f_end = _get_end_line(first)
            s_start = _get_start_line(second)
            s_end = _get_end_line(second)

            f_filename = os.path.basename(f_file)
            s_filename = os.path.basename(s_file)

            dup_lines = f_end - f_start + 1
            is_cross_file = f_filename != s_filename

            if dup_lines >= 10 and is_cross_file:
                priority = PRIORITY_HIGH
            elif dup_lines >= 6:
                priority = PRIORITY_MEDIUM
            else:
                priority = PRIORITY_LOW

            desc = (
                f"该代码块与文件 [{s_filename}] 第 {s_start}-{s_end} 行存在重复"
                f" (文件: {rel_path(project_root, s_file)})"
            )

            results.append({
                "priority": priority,
                "relative_path": rel_path(project_root, f_file),
                "filename": f_filename,
                "start_line": f_start,
                "end_line": f_end,
                "issue_type": "冗余重复代码",
                "description": desc,
            })
        except Exception as e:
            warn(f"解析 duplicate 记录异常: {e}")
            continue

    return results, errors, total_files, total_clones


# ====================================================================
# 无用代码扫描 (javalang AST)
# ====================================================================

def scan_dead_code(
    scan_target: Path,
    project_root: Path,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    使用 javalang 解析 Java 文件 AST，检测无用代码：
    - 未使用的 import
    - 未使用的 private 字段
    - 未使用的 private 方法
    - 未使用的局部变量
    """
    print("\n" + "=" * 60)
    info("无用代码扫描 - javalang AST")
    print("=" * 60)

    import javalang

    results: List[Dict[str, Any]] = []
    errors: List[str] = []

    # 收集所有 Java 文件
    java_files: List[Path] = []
    for dirpath, dirnames, filenames in os.walk(str(scan_target)):
        dirnames[:] = [d for d in dirnames if d not in EXCLUDE_DIRS]
        for f in filenames:
            if f.endswith(".java"):
                java_files.append(Path(dirpath) / f)

    info(f"找到 {len(java_files)} 个 Java 文件，开始分析...")
    parsed_count = 0
    skipped_count = 0

    for jf in java_files:
        try:
            file_size = jf.stat().st_size
        except OSError:
            skipped_count += 1
            continue

        # 大文件回退正则检测
        if file_size > 1024 * 1024:
            warn(f"文件过大，回退正则检测: {rel_path(project_root, str(jf))}")
            try:
                file_results = _dead_code_regex_only(jf, project_root)
                results.extend(file_results)
            except Exception as e:
                warn(f"正则回退检测异常: {e}")
            skipped_count += 1
            continue

        try:
            with open(jf, "r", encoding="utf-8", errors="replace") as fh:
                source = fh.read()
        except Exception as e:
            warn(f"读取失败: {rel_path(project_root, str(jf))} - {e}")
            skipped_count += 1
            continue

        try:
            tree = javalang.parse.parse(source)
        except javalang.parser.JavaSyntaxError as e:
            at_line = getattr(e.at, "line", "?") if hasattr(e, "at") else "?"
            warn(f"语法错误，跳过: {rel_path(project_root, str(jf))} at line {at_line}")
            skipped_count += 1
            continue
        except Exception as e:
            warn(f"解析异常: {rel_path(project_root, str(jf))} - {e}")
            skipped_count += 1
            continue

        parsed_count += 1
        relative = rel_path(project_root, str(jf))
        filename = jf.name
        source_lines = source.split("\n")

        try:
            _analyze_file(tree, source_lines, relative, filename, results)
        except Exception as e:
            warn(f"文件分析异常 ({relative}): {e}")
            skipped_count += 1

        if parsed_count % 100 == 0:
            info(f"已分析 {parsed_count}/{len(java_files)} 个文件...")

    if errors:
        for e in errors:
            err(e)

    info(f"分析完成: 成功 {parsed_count} 个, 跳过 {skipped_count} 个")
    info(f"发现 {len(results)} 处死代码问题")
    return results, errors


# ── 单文件 AST 分析 ────────────────────────────────────────────────

def _analyze_file(tree: Any, source_lines: List[str], relative: str,
                  filename: str, results: List[Dict[str, Any]]) -> None:
    """对单个 Java 文件执行 4 种无用代码检测。"""
    try:
        _detect_unused_imports(tree, source_lines, relative, filename, results)
    except Exception as e:
        warn(f"  import 检测异常 ({filename}): {e}")

    try:
        _detect_unused_fields(tree, source_lines, relative, filename, results)
    except Exception as e:
        warn(f"  字段检测异常 ({filename}): {e}")

    try:
        _detect_unused_methods(tree, source_lines, relative, filename, results)
    except Exception as e:
        warn(f"  方法检测异常 ({filename}): {e}")

    try:
        _detect_unused_locals(tree, source_lines, relative, filename, results)
    except Exception as e:
        warn(f"  局部变量检测异常 ({filename}): {e}")


def _detect_unused_imports(tree: Any, source_lines: List[str],
                           relative: str, filename: str,
                           results: List[Dict[str, Any]]) -> None:
    """检测未使用的 import 语句。"""
    for imp in tree.imports:
        if not imp.path:
            continue
        try:
            short_name = imp.path.split(".")[-1]
            if short_name == "*":
                pkg_prefix = imp.path.rsplit(".", 1)[0] + "."
                used = any(
                    not line.strip().startswith("import") and pkg_prefix[:-1] in line
                    for line in source_lines
                )
            else:
                used = any(
                    not line.strip().startswith("import")
                    and re.search(r"\b" + re.escape(short_name) + r"\b", line)
                    for line in source_lines
                )
        except re.error:
            continue
        except Exception:
            continue

        if not used:
            results.append({
                "priority": PRIORITY_LOW,
                "relative_path": relative,
                "filename": filename,
                "start_line": (imp.position.line if imp.position else 1),
                "end_line": (imp.position.line if imp.position else 1),
                "issue_type": "未使用的 import",
                "description": f"import '{imp.path}' 在文件中未被使用",
            })


def _detect_unused_fields(tree: Any, source_lines: List[str],
                          relative: str, filename: str,
                          results: List[Dict[str, Any]]) -> None:
    """检测未使用的 private 字段。"""
    from javalang.tree import FieldDeclaration

    private_fields: List[Tuple[Any, str]] = []
    injection_fields: Set[str] = set()
    injection_keywords = {"Autowired", "Inject", "Resource", "Value"}

    for _path, node in tree:
        try:
            if not isinstance(node, FieldDeclaration):
                continue
            modifiers = set(node.modifiers) if node.modifiers else set()
            if "private" not in modifiers:
                continue

            annotations: Set[str] = set()
            if hasattr(node, "annotations") and node.annotations:
                for ann in node.annotations:
                    if hasattr(ann, "name"):
                        annotations.add(ann.name)

            for decl in node.declarators:
                if decl.name == "serialVersionUID":
                    continue
                if annotations & injection_keywords:
                    injection_fields.add(decl.name)
                    continue
                private_fields.append((node, decl.name))
        except Exception:
            continue

    used_fields: Set[str] = set()
    for line_num, line in enumerate(source_lines, 1):
        for fnode, fn in private_fields:
            if fn in used_fields:
                continue
            try:
                if not re.search(r"\b" + re.escape(fn) + r"\b", line):
                    continue
            except re.error:
                continue
            if fnode.position and fnode.position.line == line_num:
                continue
            used_fields.add(fn)

    for fnode, fn in private_fields:
        if fn not in used_fields:
            results.append({
                "priority": PRIORITY_MEDIUM,
                "relative_path": relative,
                "filename": filename,
                "start_line": (fnode.position.line if fnode.position else 1),
                "end_line": (fnode.position.line if fnode.position else 1),
                "issue_type": "未使用的 private 字段",
                "description": f"private 字段 '{fn}' 声明后未被使用",
            })


def _detect_unused_methods(tree: Any, source_lines: List[str],
                           relative: str, filename: str,
                           results: List[Dict[str, Any]]) -> None:
    """检测未使用的 private 方法。"""
    from javalang.tree import MethodDeclaration

    private_methods: List[Tuple[Any, str]] = []
    for _path, node in tree:
        try:
            if not isinstance(node, MethodDeclaration):
                continue
            modifiers = set(node.modifiers) if node.modifiers else set()
            if "private" not in modifiers:
                continue
            if node.name == "main":
                continue
            private_methods.append((node, node.name))
        except Exception:
            continue

    used_methods: Set[str] = set()
    for line_num, line in enumerate(source_lines, 1):
        for mnode, mn in private_methods:
            if mn in used_methods:
                continue
            try:
                if not re.search(r"\b" + re.escape(mn) + r"\s*\(", line):
                    continue
            except re.error:
                continue
            if mnode.position and mnode.position.line == line_num:
                continue
            used_methods.add(mn)

    for mnode, mn in private_methods:
        if mn not in used_methods:
            results.append({
                "priority": PRIORITY_MEDIUM,
                "relative_path": relative,
                "filename": filename,
                "start_line": (mnode.position.line if mnode.position else 1),
                "end_line": (mnode.position.line if mnode.position else 1),
                "issue_type": "未使用的 private 方法",
                "description": f"private 方法 '{mn}' 定义后未被调用",
            })


def _detect_unused_locals(tree: Any, source_lines: List[str],
                          relative: str, filename: str,
                          results: List[Dict[str, Any]]) -> None:
    """检测未使用的局部变量。"""
    from javalang.tree import LocalVariableDeclaration

    local_vars: List[Tuple[str, int]] = []
    for _path, node in tree:
        try:
            if not isinstance(node, LocalVariableDeclaration):
                continue
            for decl in node.declarators:
                decl_line = node.position.line if node.position else 1
                local_vars.append((decl.name, decl_line))
        except Exception:
            continue

    used_locals: Set[str] = set()
    for line_num, line in enumerate(source_lines, 1):
        for lv_name, decl_line in local_vars:
            if lv_name in used_locals:
                continue
            if line_num <= decl_line:
                continue
            try:
                if not re.search(r"\b" + re.escape(lv_name) + r"\b", line):
                    continue
            except re.error:
                continue
            if "=" in line:
                try:
                    if re.search(r"\b" + re.escape(lv_name) + r"\s*=", line):
                        continue
                except re.error:
                    pass
            used_locals.add(lv_name)

    for lv_name, decl_line in local_vars:
        if lv_name not in used_locals:
            results.append({
                "priority": PRIORITY_LOW,
                "relative_path": relative,
                "filename": filename,
                "start_line": decl_line,
                "end_line": decl_line,
                "issue_type": "未使用的局部变量",
                "description": f"局部变量 '{lv_name}' 声明后仅赋值未读取",
            })


def _dead_code_regex_only(file_path: Path, project_root: Path) -> List[Dict[str, Any]]:
    """大文件回退: 仅用正则检测未使用的 import。"""
    results: List[Dict[str, Any]] = []
    relative = rel_path(project_root, str(file_path))
    filename = file_path.name

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
            source = fh.read()
    except Exception:
        return results

    source_lines = source.split("\n")
    import_pattern = re.compile(r"^import\s+(static\s+)?([a-zA-Z0-9_.*]+)\s*;")

    for line_num, line in enumerate(source_lines, 1):
        try:
            m = import_pattern.match(line.strip())
            if not m:
                continue
            full_path = m.group(2)
            short_name = full_path.split(".")[-1]
            if short_name == "*":
                continue

            used = any(
                i != line_num and re.search(r"\b" + re.escape(short_name) + r"\b", sl)
                for i, sl in enumerate(source_lines, 1)
            )
            if not used:
                results.append({
                    "priority": PRIORITY_LOW,
                    "relative_path": relative,
                    "filename": filename,
                    "start_line": line_num,
                    "end_line": line_num,
                    "issue_type": "未使用的 import (正则回退)",
                    "description": f"import '{full_path}' 在文件中未被使用",
                })
        except (re.error, Exception):
            continue

    return results


# ====================================================================
# 输出: Markdown 报告
# ====================================================================

def write_markdown(
    dup_results: List[Dict[str, Any]],
    dead_results: List[Dict[str, Any]],
    output_path: str,
    project_root: Path,
    project_name: str,
    min_lines: int,
    min_tokens: int,
    scan_errors: List[str],
    scan_summary: Dict[str, Any],
) -> str:
    """生成中文 Markdown 格式报告。"""
    total = len(dup_results) + len(dead_results)
    total_dup_lines = sum(
        max(0, r.get("end_line", 0) - r.get("start_line", 0) + 1)
        for r in dup_results
    )
    total_dead = len(dead_results)

    now_str = datetime.now(CST).strftime("%Y-%m-%d %H:%M CST")

    lines: List[str] = []
    lines.append(f"# Java 代码质量扫描报告")
    lines.append("")
    lines.append(f"**项目**: {project_name}")
    lines.append(f"**扫描时间**: {now_str}")
    lines.append(f"**扫描范围**: {project_root}")
    lines.append("")

    # ── 概览统计表 ──
    lines.append("## 📊 概览统计")
    lines.append("")
    lines.append("| 指标 | 数值 |")
    lines.append("|------|------|")
    lines.append(f"| 重复代码块数量 | {len(dup_results)} 个 |")
    lines.append(f"| 重复代码总行数 | {total_dup_lines} 行 |")
    lines.append(f"| 无用代码问题数 | {total_dead} 处 |")
    lines.append(f"| 问题总数 | {total} 处 |")
    lines.append(f"| jscpd 扫描文件数 | {scan_summary.get('jscpd_files', 0)} 个 |")
    lines.append(f"| javalang 分析文件数 | {scan_summary.get('dead_files', 0)} 个 |")
    lines.append(f"| 最小重复行数 | {min_lines} |")
    lines.append(f"| 最小重复 token | {min_tokens} |")
    lines.append("")

    if scan_errors:
        lines.append("### ⚠️ 扫描异常")
        lines.append("")
        for se in scan_errors:
            lines.append(f"- {se}")
        lines.append("")

    # ── 严重程度分布 ──
    high_count = sum(1 for r in dup_results if r.get("priority") == PRIORITY_HIGH)
    med_count = sum(1 for r in dup_results if r.get("priority") == PRIORITY_MEDIUM)
    low_count = sum(1 for r in dup_results if r.get("priority") == PRIORITY_LOW)

    lines.append("### 重复代码严重程度分布")
    lines.append("")
    lines.append("| 严重程度 | 数量 |")
    lines.append("|----------|------|")
    lines.append(f"| 🔴 高 | {high_count} |")
    lines.append(f"| 🟡 中 | {med_count} |")
    lines.append(f"| 🟢 低 | {low_count} |")
    lines.append("")

    if total == 0:
        lines.append("## 🎉 扫描结果")
        lines.append("")
        lines.append("太棒了！未发现任何重复代码或无用代码问题。代码质量良好。")
        lines.append("")
        lines.append("### 建议")
        lines.append("")
        lines.append("1. 继续保持良好的编码习惯")
        lines.append("2. 定期使用本工具进行代码质量检查")
        lines.append("3. 新增代码模块时建议再次扫描")
        lines.append("")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return "\n".join(lines)

    # ── Top N 重复代码（按严重程度排序） ──
    sorted_dup = sorted(dup_results, key=lambda x: (
        PRIORITY_ORDER.get(x.get("priority", PRIORITY_LOW), 99),
        -(x.get("end_line", 0) - x.get("start_line", 0) + 1),
    ))

    top_n = min(20, len(sorted_dup))
    if top_n > 0:
        lines.append(f"## 🔄 重复代码 Top {top_n}")
        lines.append("")

        for i, item in enumerate(sorted_dup[:top_n], 1):
            dup_lines_count = max(0, item.get("end_line", 0) - item.get("start_line", 0) + 1)
            emoji = {"高": "🔴", "中": "🟡", "低": "🟢"}.get(item.get("priority", ""), "")
            lines.append(f"### {i}. {emoji} {item.get('filename', '?')} (行 {item.get('start_line', 0)}-{item.get('end_line', 0)})")
            lines.append("")
            lines.append(f"- **文件**: `{item.get('relative_path', '?')}`")
            lines.append(f"- **行号**: {item.get('start_line', 0)}-{item.get('end_line', 0)}")
            lines.append(f"- **重复行数**: {dup_lines_count} 行")
            lines.append(f"- **严重程度**: {item.get('priority', '?')}")
            lines.append(f"- **详情**: {item.get('description', '')}")
            lines.append("")

    # ── 无用代码 Top ──
    if dead_results:
        sorted_dead = sorted(dead_results, key=lambda x: (
            PRIORITY_ORDER.get(x.get("priority", PRIORITY_LOW), 99),
            x.get("issue_type", ""),
        ))

        # 按类型分组统计
        type_counts: Dict[str, int] = {}
        for r in dead_results:
            it = r.get("issue_type", "其他")
            type_counts[it] = type_counts.get(it, 0) + 1

        lines.append("## 🗑️ 无用代码")
        lines.append("")

        lines.append("### 类型分布")
        lines.append("")
        lines.append("| 类型 | 数量 |")
        lines.append("|------|------|")
        for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
            lines.append(f"| {t} | {c} |")
        lines.append("")

        dead_top_n = min(15, len(sorted_dead))
        lines.append(f"### 无用代码 Top {dead_top_n}")
        lines.append("")

        for i, item in enumerate(sorted_dead[:dead_top_n], 1):
            emoji = {"未使用的 private 字段": "🔸", "未使用的 private 方法": "🔸", "未使用的 import": "🔹", "未使用的局部变量": "🔹"}.get(item.get("issue_type", ""), "")
            lines.append(f"{i}. {emoji} **{item.get('issue_type', '?')}** — `{item.get('filename', '?')}`")
            lines.append(f"   - 文件: `{item.get('relative_path', '?')}`")
            lines.append(f"   - 行号: {item.get('start_line', 0)}")
            lines.append(f"   - 说明: {item.get('description', '')}")
            lines.append("")

    # ── 重构建议 ──
    lines.append("## 💡 重构建议")
    lines.append("")

    high_items = [r for r in dup_results if r.get("priority") == PRIORITY_HIGH]
    if high_items:
        lines.append(f"1. **优先处理 {len(high_items)} 处高严重度重复代码**：这些重复代码跨文件且行数较多，建议提取公共基类或工具类")

    med_items = [r for r in dup_results if r.get("priority") == PRIORITY_MEDIUM]
    if med_items:
        lines.append(f"2. **关注 {len(med_items)} 处中等严重度重复**：在重构高优先级代码时一并处理")

    if dead_results:
        lines.append(f"3. **清理 {len(dead_results)} 处无用代码**：未使用的 import、字段、方法和变量可以安全删除，减少代码噪音")
        unused_imports = sum(1 for r in dead_results if "import" in r.get("issue_type", ""))
        if unused_imports > 0:
            lines.append(f"   - 其中 {unused_imports} 个未使用的 import 可直接使用 IDE 自动清理")

    lines.append(f"4. 重构后务必运行单元测试，确保功能不受影响")
    lines.append(f"5. 建议将本工具集成到 CI/CD 流程中，定期扫描代码质量")
    lines.append("")

    # ── 写入文件 ──
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return "\n".join(lines)


# ====================================================================
# 输出: Excel 报告
# ====================================================================

def write_excel(
    dup_results: List[Dict[str, Any]],
    dead_results: List[Dict[str, Any]],
    output_path: str,
    scan_errors: Optional[List[str]] = None,
) -> Optional[str]:
    """将扫描结果写入格式化的 Excel 文件。"""
    total = len(dup_results) + len(dead_results)
    has_errors = scan_errors and len(scan_errors) > 0

    if total == 0 and not has_errors:
        print("\n" + "=" * 60)
        info("两次扫描均未发现问题，跳过 Excel 生成")
        return None

    print("\n" + "=" * 60)
    info("聚合结果 & 生成 Excel")
    print("=" * 60)

    field_map = {
        "priority": 0,
        "relative_path": 1, "filename": 2,
        "start_line": 3, "end_line": 4,
        "issue_type": 5, "description": 6,
    }

    def sort_by_priority(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return sorted(items, key=lambda x: (
            PRIORITY_ORDER.get(x.get("priority", PRIORITY_LOW), 99),
            -(x.get("end_line", 0) - x.get("start_line", 0) + 1),
        ))

    def build_df(results: List[Dict[str, Any]], header: List[str]) -> "pd.DataFrame":
        if not results:
            return pd.DataFrame(columns=header)
        rows = []
        for item in results:
            row = {}
            for field, col_idx in field_map.items():
                row[header[col_idx]] = item.get(field, "")
            rows.append(row)
        return pd.DataFrame(rows)[header]

    try:
        dup_results = sort_by_priority(dup_results)
        dead_results = sort_by_priority(dead_results)

        df_dup = build_df(dup_results, HEADER_DUPLICATE)
        df_dead = build_df(dead_results, HEADER_DEAD_CODE)

        # 追加扫描异常
        if has_errors:
            error_rows = []
            for i, err_msg in enumerate(scan_errors):
                error_rows.append({
                    HEADER_DEAD_CODE[0]: PRIORITY_LOW,
                    HEADER_DEAD_CODE[1]: "", HEADER_DEAD_CODE[2]: "",
                    HEADER_DEAD_CODE[3]: "", HEADER_DEAD_CODE[4]: "",
                    HEADER_DEAD_CODE[5]: f"扫描异常 #{i + 1}",
                    HEADER_DEAD_CODE[6]: err_msg,
                })
            df_error = pd.DataFrame(error_rows)
            df_dead = pd.concat([df_dead, df_error], ignore_index=True) if not df_dead.empty else df_error

        abspath = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(abspath) or ".", exist_ok=True)

        with pd.ExcelWriter(abspath, engine="openpyxl") as writer:
            df_dup.to_excel(writer, sheet_name="冗余重复代码", index=False)
            df_dead.to_excel(writer, sheet_name="无用代码", index=False)

        info(f"已写入: {abspath}")
        _format_excel(abspath)
        ok("格式美化完成")

        print(f"\n  [DONE] Excel 报告: {abspath}")
        print(f"         - 冗余重复代码: {len(df_dup)} 条")
        print(f"         - 无用代码:      {len(df_dead)} 条")

        return abspath

    except Exception as e:
        err(f"Excel 写入异常: {e}")
        return None


def _format_excel(filepath: str) -> None:
    """美化 Excel 文件格式。"""
    try:
        wb = load_workbook(filepath)
        for sheet_name in ("冗余重复代码", "无用代码"):
            ws = wb[sheet_name]
            try:
                # 表头样式
                for cell in ws[1]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGN
                    cell.border = CELL_BORDER

                # 数据行样式 + 优先级行颜色
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    priority_val = str(row[0].value) if row[0].value else ""
                    row_fill = PRIORITY_FILLS.get(priority_val)
                    for cell in row:
                        cell.alignment = CELL_ALIGN
                        cell.border = CELL_BORDER
                        if row_fill:
                            cell.fill = row_fill

                # 自动列宽
                for ci in range(1, ws.max_column + 1):
                    col_letter = chr(64 + ci)
                    max_w = 10
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                                             min_col=ci, max_col=ci):
                        for cell in row:
                            if cell.value:
                                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value)[:50])
                                max_w = max(max_w, w)
                    ws.column_dimensions[col_letter].width = min(max_w + 2, 80)

                ws.freeze_panes = "A2"
            except Exception:
                warn(f"Sheet '{sheet_name}' 美化异常，跳过")
                continue
        wb.save(filepath)
    except Exception as e:
        warn(f"Excel 格式美化异常: {e}")


# ====================================================================
# 主入口
# ====================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Java Code Scanner v2 - 冗余重复代码 + 无用代码扫描",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python main.py --project-path /path/to/java/src
  python main.py --project-path ./src --output ./report.md
  python main.py --project-path ./src --min-lines 5 --min-tokens 100
  python main.py --project-path ./src --skip-jscpd
  python main.py --project-path ./src --skip-dead-code
        """,
    )
    parser.add_argument(
        "--project-path", required=True,
        help="扫描目标目录（直接传给 jscpd，智能体负责找到包含 Java 文件的目录）",
    )
    parser.add_argument(
        "--output", default=None,
        help="报告输出路径。.md 结尾生成 Markdown，.xlsx 结尾生成 Excel，否则同时生成两种",
    )
    parser.add_argument(
        "--min-lines", type=int, default=3,
        help="jscpd 最小重复行数 (默认: 3)",
    )
    parser.add_argument(
        "--min-tokens", type=int, default=50,
        help="jscpd 最小重复 token 数 (默认: 50)",
    )
    parser.add_argument(
        "--skip-jscpd", action="store_true",
        help="跳过冗余重复代码扫描",
    )
    parser.add_argument(
        "--skip-dead-code", action="store_true",
        help="跳过无用代码扫描",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  Java Code Scanner v2")
    print("  冗余重复代码 (jscpd) + 无用代码 (javalang AST)")
    print("=" * 60)

    # 1. 验证项目路径
    try:
        project_root = resolve_path(args.project_path)
        info(f"项目: {project_root}")
    except (FileNotFoundError, NotADirectoryError) as e:
        err(str(e))
        sys.exit(1)

    # 确定输出路径和格式
    output_md: Optional[str] = None
    output_xlsx: Optional[str] = None
    project_name = project_root.name or str(project_root)

    if args.output:
        out = args.output
        if out.endswith(".md"):
            output_md = out
        elif out.endswith(".xlsx"):
            output_xlsx = out
        else:
            # 同时生成两种
            output_md = out + ".md"
            output_xlsx = out + ".xlsx"
    else:
        # 默认同时生成
        output_md = os.path.join(str(project_root), "java-code-report.md")
        output_xlsx = os.path.join(str(project_root), "java-code-report.xlsx")

    info(f"输出: markdown={output_md or '(跳过)'}, excel={output_xlsx or '(跳过)'}")

    # 2. 环境检查
    env = check_environment()

    # 3. 创建临时目录
    tmpdir = tempfile.TemporaryDirectory(prefix="java-code-scanner-")
    info(f"临时目录: {tmpdir.name}\n")

    dup_results: List[Dict[str, Any]] = []
    dead_results: List[Dict[str, Any]] = []
    scan_errors: List[str] = []
    scan_summary: Dict[str, Any] = {
        "jscpd_files": 0,
        "jscpd_clones": 0,
        "dead_files": 0,
    }

    try:
        # 4a. 冗余代码扫描
        if args.skip_jscpd:
            skip("--skip-jscpd，跳过")
        elif not env.get("jscpd"):
            skip("jscpd 未安装，跳过")
        else:
            dup_results, dup_errors, jscpd_files, jscpd_clones = scan_duplicate(
                project_root, tmpdir.name, project_root,
                min_lines=args.min_lines, min_tokens=args.min_tokens,
            )
            scan_errors.extend(dup_errors)
            scan_summary["jscpd_files"] = jscpd_files
            scan_summary["jscpd_clones"] = jscpd_clones

        # 4b. 无用代码扫描
        if args.skip_dead_code:
            skip("--skip-dead-code，跳过")
        elif not env.get("javalang"):
            skip("javalang 未安装，跳过")
        else:
            dead_results, dead_errors = scan_dead_code(project_root, project_root)
            scan_errors.extend(dead_errors)
            scan_summary["dead_files"] = len(dead_results)

        # 5. 生成报告
        markdown_content: Optional[str] = None

        if output_md:
            print("\n" + "=" * 60)
            info("生成 Markdown 报告")
            print("=" * 60)
            try:
                markdown_content = write_markdown(
                    dup_results, dead_results, output_md,
                    project_root, project_name,
                    args.min_lines, args.min_tokens,
                    scan_errors, scan_summary,
                )
                ok(f"Markdown 报告: {output_md}")
            except Exception as e:
                err(f"Markdown 报告生成失败: {e}")
                scan_errors.append(f"[Markdown] {e}")

        if output_xlsx:
            try:
                if not HAS_PANDAS or not HAS_OPENPYXL:
                    skip("pandas 或 openpyxl 未安装，跳过 Excel 报告")
                else:
                    write_excel(dup_results, dead_results, output_xlsx, scan_errors)
            except Exception as e:
                err(f"Excel 报告生成失败: {e}")
                scan_errors.append(f"[Excel] {e}")

        # 6. 汇总
        total = len(dup_results) + len(dead_results)
        print(f"\n  {'=' * 58}")
        print(f"  扫描汇总: 共 {total} 处问题")
        print(f"    - 冗余重复代码: {len(dup_results)} 处")
        print(f"    - 无用代码:     {len(dead_results)} 处")
        print(f"  {'=' * 58}")

        # 打印简要的 markdown 报告到控制台
        if markdown_content:
            print("\n" + "─" * 60)
            print("  📄 Markdown 报告预览 (前 60 行)")
            print("─" * 60)
            preview_lines = markdown_content.split("\n")[:60]
            for pl in preview_lines:
                print(f"  {pl}")
            if len(markdown_content.split("\n")) > 60:
                print(f"  ... (共 {len(markdown_content.split(chr(10)))} 行)")

    except KeyboardInterrupt:
        warn("用户中断")
        sys.exit(130)
    except Exception as e:
        err(f"扫描异常: {e}")
        import traceback
        traceback.print_exc()
        # 尝试生成已有结果
        if dup_results or dead_results:
            try:
                if output_xlsx and HAS_PANDAS:
                    write_excel(dup_results, dead_results, output_xlsx)
                if output_md:
                    write_markdown(dup_results, dead_results, output_md,
                                   project_root, project_name,
                                   args.min_lines, args.min_tokens,
                                   scan_errors, scan_summary)
            except Exception:
                pass
        sys.exit(1)
    finally:
        try:
            tmpdir.cleanup()
        except Exception:
            pass

    print(f"\n  [DONE] 扫描完成！")


if __name__ == "__main__":
    main()