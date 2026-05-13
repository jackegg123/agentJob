#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Report Generator — 从 jscpd JSON 报告 + 无用代码结果生成 Excel 报告
============================================================================

生成一个 .xlsx 文件，包含以下 Sheet：
  1. 概览 — 总体统计和严重程度分布
  2. 重复代码 — 每处重复代码的详细信息
  3. 无用代码 — 每处无用代码的详细信息
  4. 重构建议 — 按优先级排列的重构建议

用法:
  python generate_excel.py --jscpd-report .jscpd-report/jscpd-report.json \
                           --dead-code-report .dead-code-results.json \
                           --scan-dir /path/to/project \
                           --output report.xlsx
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ====================================================================
# 常量
# ====================================================================

PRIORITY_HIGH = "高"
PRIORITY_MEDIUM = "中"
PRIORITY_LOW = "低"

CST = timezone(timedelta(hours=8))

# 颜色定义
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUB_TITLE_FONT = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_FONT = Font(name="微软雅黑", size=10, bold=True, color="9C0006")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
MEDIUM_FONT = Font(name="微软雅黑", size=10, bold=True, color="9C6500")
LOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOW_FONT = Font(name="微软雅黑", size=10, color="006100")

SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
CENTER_WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SEVERITY_LABEL = {PRIORITY_HIGH: "严重", PRIORITY_MEDIUM: "中等", PRIORITY_LOW: "轻微"}


# ====================================================================
# 数据解析（复用 generate_report 的逻辑）
# ====================================================================

def parse_jscpd_report(json_path: str) -> Dict[str, Any]:
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        return {"duplicates": [], "summary": {}, "error": f"文件不存在: {json_path}"}
    except json.JSONDecodeError as e:
        return {"duplicates": [], "summary": {}, "error": f"JSON 解析失败: {e}"}

    stats = data.get("statistics", {}).get("total", {})
    summary = {
        "scanned_files": stats.get("sources", 0),
        "total_clones": stats.get("clones", 0),
        "total_dup_lines": stats.get("duplicatedLines", 0),
        "percentage": stats.get("percentage", 0),
    }

    duplicates: List[Dict[str, Any]] = []
    for dup in data.get("duplicates", []):
        try:
            first = dup.get("firstFile", {})
            second = dup.get("secondFile", {})

            f_file = first.get("name", "") or first.get("path", "")
            s_file = second.get("name", "") or second.get("path", "")
            if not f_file or not s_file:
                continue

            f_start = first.get("start", 0)
            f_end = first.get("end", 0)
            if not isinstance(f_start, int):
                f_start = first.get("startLoc", {}).get("line", 0)
            if not isinstance(f_end, int):
                f_end = first.get("endLoc", {}).get("line", 0)

            s_start = second.get("start", 0)
            s_end = second.get("end", 0)
            if not isinstance(s_start, int):
                s_start = second.get("startLoc", {}).get("line", 0)
            if not isinstance(s_end, int):
                s_end = second.get("endLoc", {}).get("line", 0)

            f_filename = os.path.basename(f_file)
            s_filename = os.path.basename(s_file)
            dup_lines = f_end - f_start + 1
            is_cross_file = f_filename != s_filename

            if dup_lines >= 20 and is_cross_file:
                priority = PRIORITY_HIGH
            elif dup_lines >= 10:
                priority = PRIORITY_MEDIUM
            else:
                priority = PRIORITY_LOW

            duplicates.append({
                "priority": priority,
                "first_file_rel": f_file,
                "first_filename": f_filename,
                "first_start": f_start,
                "first_end": f_end,
                "second_file_rel": s_file,
                "second_filename": s_filename,
                "second_start": s_start,
                "second_end": s_end,
                "dup_lines": dup_lines,
                "is_cross_file": is_cross_file,
                "fragment": dup.get("fragment", ""),
            })
        except Exception:
            continue

    ord_map = {PRIORITY_HIGH: 0, PRIORITY_MEDIUM: 1, PRIORITY_LOW: 2}
    duplicates.sort(key=lambda x: (ord_map.get(x["priority"], 99), -x["dup_lines"]))
    return {"duplicates": duplicates, "summary": summary, "error": None}


def parse_dead_code_report(json_path: str) -> Dict[str, Any]:
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {"results": data.get("results", []), "summary": data.get("summary", {}), "error": None}
    except FileNotFoundError:
        return {"results": [], "summary": {}, "error": None}
    except (json.JSONDecodeError, Exception) as e:
        return {"results": [], "summary": {}, "error": str(e)}


# ====================================================================
# 样式辅助函数
# ====================================================================

def _apply_style(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        _apply_style(ws, row, col_start + i, h, font=HEADER_FONT, fill=HEADER_FILL,
                     alignment=CENTER_WRAP_ALIGN, border=THIN_BORDER)


def _write_data_row(ws, row, data, col_start=1, alt=False):
    fill = LIGHT_GRAY_FILL if alt else None
    for i, val in enumerate(data):
        _apply_style(ws, row, col_start + i, val, font=NORMAL_FONT, fill=fill,
                     alignment=WRAP_ALIGN, border=THIN_BORDER)


def _write_priority_row(ws, row, data, priority, col_start=1, alt=False):
    """写入带颜色的行（根据严重程度着色）"""
    if priority == PRIORITY_HIGH:
        fill = HIGH_FILL
        font = HIGH_FONT
    elif priority == PRIORITY_MEDIUM:
        fill = MEDIUM_FILL
        font = MEDIUM_FONT
    elif priority == PRIORITY_LOW:
        fill = LOW_FILL
        font = LOW_FONT
    else:
        fill = LIGHT_GRAY_FILL if alt else None
        font = NORMAL_FONT

    for i, val in enumerate(data):
        _apply_style(ws, row, col_start + i, val, font=font, fill=fill,
                     alignment=WRAP_ALIGN, border=THIN_BORDER)


def _auto_width(ws, min_width=8, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                # 计算中文字符宽度（约2倍）
                val = str(cell.value)
                lines = val.split("\n")
                for line in lines:
                    length = 0
                    for ch in line:
                        length += 2 if ord(ch) > 127 else 1
                    max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ====================================================================
# Sheet 1: 概览
# ====================================================================

def _write_overview_sheet(ws, dup_data, dead_data, scan_dir=None, project_name=None):
    duplicates = dup_data.get("duplicates", [])
    dup_summary = dup_data.get("summary", {})
    dead_results = dead_data.get("results", [])
    dead_summary = dead_data.get("summary", {})

    total_dup = len(duplicates)
    total_dead = len(dead_results)
    total = total_dup + total_dead

    now_str = datetime.now(CST).strftime("%Y-%m-%d %H:%M CST")
    proj = project_name or os.path.basename(scan_dir or os.getcwd())
    scan_range = scan_dir or os.getcwd()

    # 标题
    _apply_style(ws, 1, 1, "Java 代码质量分析报告", font=TITLE_FONT)
    ws.merge_cells("A1:D1")

    _apply_style(ws, 2, 1, f"项目: {proj}", font=BOLD_FONT)
    ws.merge_cells("A2:D2")
    _apply_style(ws, 3, 1, f"扫描时间: {now_str}", font=NORMAL_FONT)
    ws.merge_cells("A3:D3")
    _apply_style(ws, 4, 1, f"扫描范围: {scan_range}", font=NORMAL_FONT)
    ws.merge_cells("A4:D4")

    # 总览统计
    row = 6
    _apply_style(ws, row, 1, "📊 概览统计", font=SUB_TITLE_FONT)
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    _write_header_row(ws, row, ["指标", "数值", "说明", ""])
    row += 1

    stats = [
        ("重复代码块数量", f"{total_dup} 个", "检测到的重复代码块数"),
        ("重复代码总行数", f"{dup_summary.get('total_dup_lines', 0)} 行", "重复代码的累计行数"),
        ("jscpd 扫描文件数", f"{dup_summary.get('scanned_files', 0)} 个", "jscpd 扫描的 Java 文件数"),
        ("jscpd 重复率", f"{dup_summary.get('percentage', 0):.2f}%", "重复行数 / 总行数"),
        ("无用代码问题数", f"{total_dead} 处", "检测到的无用代码问题"),
        ("javalang 分析文件数", f"{dead_summary.get('scanned_files', 0)} 个", "AST 解析的文件数"),
        ("问题总数", f"{total} 处", ""),
    ]
    for i, (label, value, note) in enumerate(stats):
        _write_data_row(ws, row, [label, value, note, ""], alt=(i % 2 == 1))
        row += 1

    # 严重程度分布
    if total_dup > 0:
        row += 1
        _apply_style(ws, row, 1, "重复代码严重程度分布", font=SUB_TITLE_FONT)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        _write_header_row(ws, row, ["严重程度", "数量", "说明", ""])
        row += 1

        high_count = sum(1 for d in duplicates if d.get("priority") == PRIORITY_HIGH)
        med_count = sum(1 for d in duplicates if d.get("priority") == PRIORITY_MEDIUM)
        low_count = sum(1 for d in duplicates if d.get("priority") == PRIORITY_LOW)

        sev_data = [
            ("🔴 高", high_count, "≥20行 且跨文件", PRIORITY_HIGH),
            ("🟡 中", med_count, "≥10行", PRIORITY_MEDIUM),
            ("🟢 低", low_count, "<10行 或同文件内", PRIORITY_LOW),
        ]
        for i, (label, cnt, desc, prio) in enumerate(sev_data):
            if cnt > 0:
                _write_priority_row(ws, row, [label, cnt, desc, ""], prio, alt=(i % 2 == 1))
                row += 1

    # 无用代码类型分布
    if total_dead > 0:
        row += 1
        _apply_style(ws, row, 1, "无用代码类型分布", font=SUB_TITLE_FONT)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        _write_header_row(ws, row, ["问题类型", "数量", "", ""])
        row += 1

        type_counts: Dict[str, int] = {}
        for r in dead_results:
            it = r.get("issue_type", "其他")
            type_counts[it] = type_counts.get(it, 0) + 1

        for i, (t, c) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1])):
            _write_data_row(ws, row, [t, c, "", ""], alt=(i % 2 == 1))
            row += 1

    _auto_width(ws)


# ====================================================================
# Sheet 2: 重复代码详情
# ====================================================================

def _write_duplicates_sheet(ws, dup_data):
    duplicates = dup_data.get("duplicates", [])
    dup_error = dup_data.get("error")

    _apply_style(ws, 1, 1, "🔄 重复代码详情", font=TITLE_FONT)
    ws.merge_cells("A1:I1")

    if dup_error:
        _apply_style(ws, 2, 1, f"⚠️ jscpd 扫描异常: {dup_error}", font=NORMAL_FONT,
                      fill=HIGH_FILL)
        ws.merge_cells("A2:I2")
        return

    total = len(duplicates)
    _apply_style(ws, 2, 1, f"共检测到 {total} 处重复代码", font=NORMAL_FONT)
    ws.merge_cells("A2:I2")

    headers = [
        "序号", "严重程度", "文件 A (相对路径)", "文件 A 文件名", "起始行 A",
        "结束行 A", "文件 B (相对路径)", "文件 B 文件名", "起始行 B",
        "结束行 B", "重复行数", "是否跨文件", "重复代码片段"
    ]
    _write_header_row(ws, 3, headers)
    ws.freeze_panes = "A4"

    row = 4
    for i, d in enumerate(duplicates, 1):
        cross = "是" if d.get("is_cross_file") else "否"
        frag = d.get("fragment", "").strip()
        if len(frag) > 500:
            frag = frag[:500] + "\n... (省略)"

        data = [
            i, d["priority"], d["first_file_rel"], d["first_filename"],
            d["first_start"], d["first_end"],
            d["second_file_rel"], d["second_filename"],
            d["second_start"], d["second_end"],
            d["dup_lines"], cross, frag
        ]
        _write_priority_row(ws, row, data, d["priority"], alt=(i % 2 == 1))
        row += 1

    _auto_width(ws, max_width=50)
    # 代码片段列设宽一些
    ws.column_dimensions[get_column_letter(13)].width = 50


# ====================================================================
# Sheet 3: 无用代码详情
# ====================================================================

def _write_deadcode_sheet(ws, dead_data):
    dead_results = dead_data.get("results", [])
    dead_error = dead_data.get("error")

    _apply_style(ws, 1, 1, "🗑️ 无用代码详情", font=TITLE_FONT)
    ws.merge_cells("A1:G1")

    if dead_error:
        _apply_style(ws, 2, 1, f"⚠️ 无用代码扫描异常: {dead_error}", font=NORMAL_FONT,
                      fill=HIGH_FILL)
        ws.merge_cells("A2:G2")
        return

    total = len(dead_results)
    _apply_style(ws, 2, 1, f"共检测到 {total} 处无用代码问题", font=NORMAL_FONT)
    ws.merge_cells("A2:G2")

    headers = [
        "序号", "严重程度", "问题类型", "文件相对路径", "文件名",
        "起始行", "结束行", "详细说明"
    ]
    _write_header_row(ws, 3, headers)
    ws.freeze_panes = "A4"

    ord_map = {PRIORITY_HIGH: 0, PRIORITY_MEDIUM: 1, PRIORITY_LOW: 2}
    sorted_dead = sorted(dead_results, key=lambda x: (
        ord_map.get(x.get("priority", PRIORITY_LOW), 99),
        x.get("issue_type", ""),
    ))

    row = 4
    for i, r in enumerate(sorted_dead, 1):
        data = [
            i,
            r.get("priority", ""),
            r.get("issue_type", ""),
            r.get("relative_path", ""),
            r.get("filename", ""),
            r.get("start_line", 0),
            r.get("end_line", 0),
            r.get("description", ""),
        ]
        _write_priority_row(ws, row, data, r.get("priority", ""), alt=(i % 2 == 1))
        row += 1

    _auto_width(ws, max_width=50)


# ====================================================================
# Sheet 4: 重构建议
# ====================================================================

def _write_suggestions_sheet(ws, dup_data, dead_data):
    duplicates = dup_data.get("duplicates", [])
    dead_results = dead_data.get("results", [])

    high_count = sum(1 for d in duplicates if d.get("priority") == PRIORITY_HIGH)
    med_count = sum(1 for d in duplicates if d.get("priority") == PRIORITY_MEDIUM)
    total_dead = len(dead_results)

    _apply_style(ws, 1, 1, "💡 重构建议", font=TITLE_FONT)
    ws.merge_cells("A1:C1")

    headers = ["优先级", "建议编号", "重构建议", "详细说明"]
    _write_header_row(ws, 2, headers)

    suggestions: List[Tuple[str, int, str, str]] = []

    n = 1
    if high_count > 0:
        suggestions.append((
            "🔴 高", n,
            "优先处理高严重度重复代码",
            f"存在 {high_count} 处高严重度重复代码（跨文件且≥20行），建议提取公共基类、工具类或模板方法来消除重复",
        ))
        n += 1

    if med_count > 0:
        suffix = "在重构高优先级代码时一并处理" if high_count > 0 else "建议逐步重构"
        suggestions.append((
            "🟡 中", n,
            "关注中等严重度重复代码",
            f"存在 {med_count} 处中等严重度重复代码，{suffix}",
        ))
        n += 1

    if total_dead > 0:
        unused_imports = sum(1 for r in dead_results if "import" in r.get("issue_type", ""))
        suggestions.append((
            "🔵 清理", n,
            "清理无用代码",
            f"存在 {total_dead} 处无用代码（其中 {unused_imports} 个未使用的 import 可 IDE 自动清理），删除后可减少代码噪音",
        ))
        n += 1

    # 通用建议
    suggestions.append((
        "🔵 通用", n,
        "重构后运行单元测试",
        "确保所有修改不影响现有功能，建议 CI/CD 中集成自动化测试",
    ))
    n += 1

    suggestions.append((
        "🔵 通用", n,
        "集成到 CI/CD 流程",
        "将本扫描工具集成到 CI/CD 流程中，定期执行代码质量检查，提前发现问题",
    ))

    row = 3
    for i, (prio, num, title, detail) in enumerate(suggestions):
        fill = HIGH_FILL if prio.startswith("🔴") else (
            MEDIUM_FILL if prio.startswith("🟡") else LIGHT_GRAY_FILL if i % 2 == 1 else None
        )
        font = HIGH_FONT if prio.startswith("🔴") else (
            MEDIUM_FONT if prio.startswith("🟡") else NORMAL_FONT
        )
        _apply_style(ws, row, 1, prio, font=font, fill=fill, alignment=CENTER_ALIGN, border=THIN_BORDER)
        _apply_style(ws, row, 2, num, font=NORMAL_FONT, fill=fill, alignment=CENTER_ALIGN, border=THIN_BORDER)
        _apply_style(ws, row, 3, title, font=BOLD_FONT, fill=fill, alignment=WRAP_ALIGN, border=THIN_BORDER)
        _apply_style(ws, row, 4, detail, font=NORMAL_FONT, fill=fill, alignment=WRAP_ALIGN, border=THIN_BORDER)
        row += 1

    _auto_width(ws)
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 60


# ====================================================================
# 主函数
# ====================================================================

def generate_excel(
    dup_data: Dict[str, Any],
    dead_data: Dict[str, Any],
    scan_dir: Optional[str] = None,
    project_name: Optional[str] = None,
    output_path: str = "java-analysis-report.xlsx",
) -> str:
    """生成 Excel 报告。"""
    wb = openpyxl.Workbook()

    # Sheet 1: 概览
    ws_overview = wb.active
    ws_overview.title = "概览"
    _write_overview_sheet(ws_overview, dup_data, dead_data,
                           scan_dir=scan_dir, project_name=project_name)

    # Sheet 2: 重复代码
    ws_dup = wb.create_sheet("重复代码")
    _write_duplicates_sheet(ws_dup, dup_data)

    # Sheet 3: 无用代码
    ws_dead = wb.create_sheet("无用代码")
    _write_deadcode_sheet(ws_dead, dead_data)

    # Sheet 4: 重构建议
    ws_sug = wb.create_sheet("重构建议")
    _write_suggestions_sheet(ws_sug, dup_data, dead_data)

    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从 jscpd + 无用代码结果生成 Excel 报告",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python generate_excel.py --jscpd-report .jscpd-report/jscpd-report.json \\
                            --dead-code-report .dead-code-results.json \\
                            --scan-dir ./src -o report.xlsx
        """,
    )
    parser.add_argument("--jscpd-report", default=None,
                        help="jscpd JSON 报告路径")
    parser.add_argument("--dead-code-report", default=None,
                        help="无用代码扫描 JSON 结果路径")
    parser.add_argument("--scan-dir", default=None,
                        help="扫描目录（用于报告头部）")
    parser.add_argument("--project-name", default=None,
                        help="项目名称")
    parser.add_argument("--output", "-o", default="java-analysis-report.xlsx",
                        help="输出 Excel 文件路径")
    args = parser.parse_args()

    # 解析报告
    dup_data = {"duplicates": [], "summary": {}, "error": None}
    if args.jscpd_report:
        if not os.path.exists(args.jscpd_report):
            dup_data["error"] = f"jscpd 报告文件不存在: {args.jscpd_report}"
        else:
            dup_data = parse_jscpd_report(args.jscpd_report)
    else:
        dup_data["error"] = "未提供 jscpd 报告（--jscpd-report）"

    dead_data = {"results": [], "summary": {}, "error": None}
    if args.dead_code_report and os.path.exists(args.dead_code_report):
        dead_data = parse_dead_code_report(args.dead_code_report)

    output_path = generate_excel(
        dup_data, dead_data,
        scan_dir=args.scan_dir,
        project_name=args.project_name,
        output_path=args.output,
    )

    dup_count = len(dup_data.get("duplicates", []))
    dead_count = len(dead_data.get("results", []))
    print(f"✅ Excel 报告已生成: {output_path}")
    print(f"   - 重复代码: {dup_count} 处")
    print(f"   - 无用代码: {dead_count} 处")
    print(f"   - 总计: {dup_count + dead_count} 处")


if __name__ == "__main__":
    main()
