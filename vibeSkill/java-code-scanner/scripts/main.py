#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Java Code Scanner Skill - 核心业务逻辑
============================================
两大核心能力：
  1. 冗余重复代码扫描（基于 jscpd）
  2. 无用代码 / Dead Code 扫描（基于 Qodana JVM Community）

用法:
  python main.py --project-path /path/to/java/project [--output /path/to/report.xlsx]

环境变量控制:
  SKIP_JSCPD=1  跳过冗余代码扫描
  SKIP_QODANA=1  跳过无用代码扫描
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ====================================================================
# 常量定义
# ====================================================================

# Sheet 1 表头：冗余重复代码
HEADER_DUPLICATE = [
    "项目相对路径",
    "文件名",
    "起始行",
    "结束行",
    "问题类型",
    "详细说明（重复代码位置）",
]

# Sheet 2 表头：无用代码
HEADER_DEAD_CODE = [
    "项目相对路径",
    "文件名",
    "起始行",
    "结束行",
    "问题类型",
    "详细说明（为什么无用）",
]

# 需要过滤的 Qodana 规则关键词
QODANA_DEAD_CODE_RULES = ["UnusedDeclaration", "unused", "DeadCode", "Unused"]

# Excel 样式常量
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ====================================================================
# 工具函数
# ====================================================================

def resolve_project_path(project_path: str) -> Path:
    """
    规范化并验证项目路径。

    Args:
        project_path: 用户传入的项目路径字符串。

    Returns:
        规范化后的 Path 对象。

    Raises:
        FileNotFoundError: 路径不存在。
        NotADirectoryError: 路径不是目录。
    """
    path = Path(project_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"项目路径不存在: {project_path}")
    if not path.is_dir():
        raise NotADirectoryError(f"项目路径不是目录: {project_path}")
    return path


def to_relative_path(project_root: Path, absolute_path: str) -> str:
    """
    将绝对文件路径转为项目相对路径。

    Args:
        project_root: 项目根目录。
        absolute_path: 文件的绝对路径。

    Returns:
        相对路径字符串。
    """
    try:
        return str(Path(absolute_path).resolve().relative_to(project_root))
    except (ValueError, RuntimeError):
        return Path(absolute_path).name


def run_cmd(cmd: List[str], cwd: Optional[str] = None, timeout: int = 1800) -> subprocess.CompletedProcess:
    """
    安全执行外部命令，统一处理超时与异常。

    Args:
        cmd: 命令列表。
        cwd: 工作目录。
        timeout: 超时秒数，默认 30 分钟。

    Returns:
        subprocess.CompletedProcess 对象。
    """
    print(f"  ▶ {' '.join(cmd)}")
    try:
        result = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True, timeout=timeout,
        )
        if result.returncode != 0:
            print(f"  ⚠️  退出码 {result.returncode}")
            if result.stderr:
                stderr_preview = result.stderr.strip()[:500]
                print(f"  └─ stderr: {stderr_preview}")
        return result
    except subprocess.TimeoutExpired:
        print(f"  ❌ 命令超时 ({timeout}s): {' '.join(cmd)}")
        raise
    except FileNotFoundError:
        print(f"  ❌ 命令未找到，请确认已安装: {cmd[0]}")
        raise
    except Exception as e:
        print(f"  ❌ 执行失败: {e}")
        raise


# ====================================================================
# 模块 2：jscpd 冗余代码扫描
# ====================================================================

def scan_duplicate(project_root: Path, temp_dir: str) -> List[Dict[str, Any]]:
    """
    使用 jscpd 扫描 Java 项目中的重复代码。

    Args:
        project_root: 项目根目录。
        temp_dir: 临时目录，存放 jscpd JSON 报告。

    Returns:
        扫描结果列表，每个字典包含：
            relative_path / filename / start_line / end_line / issue_type / description。
    """
    print("\n═══════════════════════════════════════════════")
    print("  [模块 1/2] 冗余重复代码扫描 — jscpd")
    print("═══════════════════════════════════════════════")

    results: List[Dict[str, Any]] = []

    if not shutil.which("jscpd") and not shutil.which("npx"):
        print("  ⏭️  跳过：jscpd / npx 未安装。")
        return results

    try:
        cmd = [
            "npx", "jscpd",
            str(project_root),
            "--pattern", "**/*.java",
            "--reporters", "json",
            "--output", temp_dir,
            "--threshold", "0",
            "--min-lines", "5",
            "--min-tokens", "50",
        ]
        run_cmd(cmd, cwd=str(project_root), timeout=600)

        # 查找生成的报告文件（jscpd 输出路径不确定，需遍历）
        report_path = None
        for root, _, files in os.walk(temp_dir):
            for f in files:
                if f.endswith(".json"):
                    report_path = os.path.join(root, f)
                    break
            if report_path:
                break

        if not report_path:
            print("  ℹ️  jscpd 未生成报告，未发现重复代码。")
            return results

        print(f"  📄 解析报告: {report_path}")
        with open(report_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        duplicates = data.get("duplicates", [])
        print(f"  🔍 发现 {len(duplicates)} 处重复")

        for dup in duplicates:
            first = dup.get("first", {})
            second = dup.get("second", {})

            def safe_line(loc):
                """统一处理行号（可能是 int 或 {"line": N}）"""
                if isinstance(loc, dict):
                    return loc.get("line", 0)
                return loc or 0

            f_file = first.get("name", "") or first.get("path", "")
            f_start = safe_line(first.get("start", first.get("startLine", 0)))
            f_end = safe_line(first.get("end", first.get("endLine", 0)))

            s_file = second.get("name", "") or second.get("path", "")
            s_start = safe_line(second.get("start", second.get("startLine", 0)))
            s_end = safe_line(second.get("end", second.get("endLine", 0)))

            s_filename = os.path.basename(s_file)
            desc = (
                f"该代码块与文件「{s_filename}」第 {s_start}-{s_end} 行存在重复"
                f"（文件: {to_relative_path(project_root, s_file)}）"
            )

            results.append({
                "relative_path": to_relative_path(project_root, f_file),
                "filename": os.path.basename(f_file),
                "start_line": f_start,
                "end_line": f_end,
                "issue_type": "冗余重复代码",
                "description": desc,
            })

        return results

    except Exception as e:
        print(f"  ❌ jscpd 扫描异常: {e}")
        return results


# ====================================================================
# 模块 3：Qodana 无用代码扫描
# ====================================================================

def scan_dead_code(project_root: Path, temp_dir: str) -> List[Dict[str, Any]]:
    """
    使用 Qodana 扫描 Java 项目中的无用代码。

    解析 SARIF 报告，过滤 UnusedDeclaration / DeadCode 等规则。

    Args:
        project_root: 项目根目录。
        temp_dir: 临时目录，存放 Qodana 输出。

    Returns:
        扫描结果列表。
    """
    print("\n═══════════════════════════════════════════════")
    print("  [模块 2/2] 无用代码扫描 — Qodana JVM Community")
    print("═══════════════════════════════════════════════")

    results: List[Dict[str, Any]] = []
    results_dir = os.path.join(temp_dir, "qodana-results")
    os.makedirs(results_dir, exist_ok=True)

    if not shutil.which("qodana"):
        print("  ⏭️  跳过：qodana CLI 未安装。")
        return results

    try:
        cmd = [
            "qodana", "scan",
            "-s", str(project_root),
            "-o", results_dir,
            "--profile-name", "qodana.recommended",
            "--linter", "jetbrains/qodana-jvm-community",
        ]
        run_cmd(cmd, timeout=3600)

        # 查找 SARIF 报告
        sarif_paths = []
        for root, _, files in os.walk(results_dir):
            for f in files:
                if f.endswith(".sarif.json") or (f == "qodana.sarif.json"):
                    sarif_paths.append(os.path.join(root, f))

        if not sarif_paths:
            # 兜底：尝试临时目录下的文件名匹配
            fallback = os.path.join(results_dir, "qodana.sarif.json")
            if os.path.exists(fallback):
                sarif_paths.append(fallback)

        if not sarif_paths:
            print("  ℹ️  Qodana 未生成 SARIF 报告。")
            return results

        sarif_path = sarif_paths[0]
        print(f"  📄 解析报告: {sarif_path}")

        with open(sarif_path, "r", encoding="utf-8", errors="replace") as f:
            sarif = json.load(f)

        for run in sarif.get("runs", []):
            # 构建 artifact 索引
            artifacts = run.get("artifacts", [])
            artifact_map: Dict[int, str] = {}
            for idx, art in enumerate(artifacts):
                loc = art.get("location", {})
                artifact_map[idx] = loc.get("uri", "")

            results_list = run.get("results", [])
            for res in results_list:
                rule_id = res.get("ruleId", "")
                msg = res.get("message", {}).get("text", "")

                # 过滤无用代码相关的规则
                is_dead = any(k.lower() in rule_id.lower() for k in QODANA_DEAD_CODE_RULES)
                if not is_dead:
                    ml = msg.lower()
                    is_dead = any(k in ml for k in ["unused", "dead code", "never used", "not used"])
                if not is_dead:
                    continue

                locations = res.get("locations", [])
                if not locations:
                    continue

                phys = locations[0].get("physicalLocation", {})
                art_loc = phys.get("artifactLocation", {})
                region = phys.get("region", {})

                uri = art_loc.get("uri", "")
                uri_idx = art_loc.get("index")
                if not uri and uri_idx is not None:
                    uri = artifact_map.get(uri_idx, "")

                start_line = region.get("startLine", 0)
                end_line = region.get("endLine", start_line)

                results.append({
                    "relative_path": to_relative_path(project_root, uri),
                    "filename": os.path.basename(uri),
                    "start_line": start_line,
                    "end_line": end_line,
                    "issue_type": rule_id or "UnusedCode",
                    "description": msg.strip() if msg else f"检测到 {rule_id}",
                })

        print(f"  🔍 发现 {len(results)} 处无用代码问题")
        return results

    except subprocess.TimeoutExpired:
        print("  ❌ Qodana 扫描超时。")
        return results
    except Exception as e:
        print(f"  ❌ Qodana 扫描异常: {e}")
        return results


# ====================================================================
# 模块 4：结果聚合与 Excel 生成
# ====================================================================

def write_excel(
    dup_results: List[Dict[str, Any]],
    dead_results: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    将扫描结果写入格式化的 Excel 文件。

    两个 Sheet：「冗余重复代码」和「无用代码」。

    Args:
        dup_results: 重复代码扫描结果。
        dead_results: 无用代码扫描结果。
        output_path: 输出路径。

    Returns:
        最终输出的文件路径。
    """
    print("\n═══════════════════════════════════════════════")
    print("  [输出] 聚合结果 & 生成 Excel")
    print("═══════════════════════════════════════════════")

    # 构建 DataFrame
    df_dup = pd.DataFrame(dup_results) if dup_results else pd.DataFrame(columns=HEADER_DUPLICATE)
    if not df_dup.empty:
        df_dup = df_dup.rename(columns={
            "relative_path": HEADER_DUPLICATE[0],
            "filename": HEADER_DUPLICATE[1],
            "start_line": HEADER_DUPLICATE[2],
            "end_line": HEADER_DUPLICATE[3],
            "issue_type": HEADER_DUPLICATE[4],
            "description": HEADER_DUPLICATE[5],
        })[HEADER_DUPLICATE]

    df_dead = pd.DataFrame(dead_results) if dead_results else pd.DataFrame(columns=HEADER_DEAD_CODE)
    if not df_dead.empty:
        df_dead = df_dead.rename(columns={
            "relative_path": HEADER_DEAD_CODE[0],
            "filename": HEADER_DEAD_CODE[1],
            "start_line": HEADER_DEAD_CODE[2],
            "end_line": HEADER_DEAD_CODE[3],
            "issue_type": HEADER_DEAD_CODE[4],
            "description": HEADER_DEAD_CODE[5],
        })[HEADER_DEAD_CODE]

    abspath = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(abspath) or ".", exist_ok=True)

    with pd.ExcelWriter(abspath, engine="openpyxl") as writer:
        df_dup.to_excel(writer, sheet_name="冗余重复代码", index=False)
        df_dead.to_excel(writer, sheet_name="无用代码", index=False)

    print(f"  💾 已写入: {abspath}")

    # 美化格式
    wb = load_workbook(abspath)
    for sheet_name in ("冗余重复代码", "无用代码"):
        ws = wb[sheet_name]

        # 表头样式
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # 数据单元格样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        # 自适应列宽
        for col_idx in range(1, ws.max_column + 1):
            col_letter = chr(64 + col_idx)
            max_w = 10
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_w = max(max_w, w)
            ws.column_dimensions[col_letter].width = min(max_w + 2, 80)

        ws.freeze_panes = "A2"

    wb.save(abspath)
    print("  🎨 格式美化完成。")
    print(f"\n  ✅ 报告: {abspath}")
    print(f"     「冗余重复代码」: {len(df_dup)} 条")
    print(f"     「无用代码」:      {len(df_dead)} 条")

    return abspath


# ====================================================================
# 主入口
# ====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Java Code Scanner — 冗余代码 + 无用代码扫描工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python main.py --project-path /path/to/java/project
  python main.py --project-path /path --output ./report.xlsx
  SKIP_JSCPD=1 python main.py --project-path /path
        """,
    )
    parser.add_argument("--project-path", required=True, help="待扫描的 Java 项目根目录")
    parser.add_argument("--output", default=None, help="Excel 报告输出路径")
    args = parser.parse_args()

    print("🍄 Java Code Scanner Skill v1.0.0")
    print()

    # 1. 验证项目路径
    try:
        project_root = resolve_project_path(args.project_path)
        print(f"  📂 项目: {project_root}")
    except (FileNotFoundError, NotADirectoryError) as e:
        print(f"  ❌ {e}")
        sys.exit(1)

    output_path = args.output or os.path.join(str(project_root), "java-code-report.xlsx")
    print(f"  📄 输出: {output_path}")
    print(f"  🕐 {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # 2. 创建临时目录
    tmpdir = tempfile.TemporaryDirectory(prefix="java-code-scanner-")
    print(f"  📁 临时: {tmpdir.name}\n")

    dup_results: List[Dict[str, Any]] = []
    dead_results: List[Dict[str, Any]] = []

    try:
        # 3. 冗余代码扫描
        if os.environ.get("SKIP_JSCPD", "").lower() in ("1", "true", "yes"):
            print("  ⏭️  SKIP_JSCPD=1，跳过。")
        else:
            dup_results = scan_duplicate(project_root, tmpdir.name)

        # 4. 无用代码扫描
        if os.environ.get("SKIP_QODANA", "").lower() in ("1", "true", "yes"):
            print("  ⏭️  SKIP_QODANA=1，跳过。")
        else:
            dead_results = scan_dead_code(project_root, tmpdir.name)

        # 5. 生成 Excel
        write_excel(dup_results, dead_results, output_path)

        # 6. 汇总
        total = len(dup_results) + len(dead_results)
        print(f"\n  📊 共计 {total} 处问题")
        if total == 0:
            print("  🎉 代码质量不错，没有发现问题！")

    except KeyboardInterrupt:
        print("\n  ⛔ 用户中断。")
        sys.exit(130)
    except Exception as e:
        print(f"\n  💥 异常: {e}")
        # 尝试保存已有结果
        try:
            write_excel(dup_results, dead_results, output_path)
        except Exception:
            pass
        sys.exit(1)
    finally:
        try:
            tmpdir.cleanup()
        except Exception:
            pass

    print("  🎉 扫描完成！")


if __name__ == "__main__":
    main()
